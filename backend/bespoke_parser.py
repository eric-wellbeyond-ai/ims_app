"""
Bespoke parser for the MPFM Validation Data spreadsheet format.
===============================================================
Anything that encodes knowledge of a *specific* file layout — hard-coded row
indices, column letters, sheet names, or company-specific column ordering —
lives here and nowhere else.

The generic analysis pipeline (mpfm_analysis.py) accepts plain DataFrames and
has no knowledge of this file format.

CLI usage (from the ims_app/ directory):
    python -m backend.bespoke_parser <spreadsheet.xlsx> [--sheet SHEET] [--plots] [--csv]
"""

import argparse
from pathlib import Path

import pandas as pd

from backend.mpfm_analysis import (
    PVTProperties,
    TestWindow,
    MPFMAnalysisResult,
    MeterAggregation,
    filter_test_window,
    compute_derived_columns,
    compute_deviations,
    build_comparison_table,
)


# ---------------------------------------------------------------------------
# Cell / range helpers
# ---------------------------------------------------------------------------

def _cell(ws, row: int, col: int):
    """Read a single cell value from an openpyxl worksheet (1-indexed)."""
    return ws.cell(row=row, column=col).value


# ---------------------------------------------------------------------------
# Metadata reader  (knows rows 9 and 12, columns L / N / O / D)
# ---------------------------------------------------------------------------

def read_metadata(filepath: str, sheet_name: str = "MPFM VALIDATION DATA"):
    """
    Extract PVT properties and test window from the spreadsheet header.

    Hard-coded layout:
      Row 9, col 14 (N9) → oil shrinkage
      Row 9, col 15 (O9) → flash factor (scf/stb)
      Row 9, col 12 (L9) → BS&W fraction
      Row 12, col 1  (A12) → first data row index
      Row 12, col 2  (B12) → last  data row index
      Column 4 (D)         → timestamps

    Returns (PVTProperties, TestWindow).
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    oil_shrinkage = float(_cell(ws, 9, 14))   # N9
    flash_factor  = float(_cell(ws, 9, 15))   # O9
    bsw           = float(_cell(ws, 9, 12))   # L9

    start_row_idx = int(_cell(ws, 12, 1))     # A12
    end_row_idx   = int(_cell(ws, 12, 2))     # B12
    start_time = pd.Timestamp(_cell(ws, start_row_idx, 4))
    end_time   = pd.Timestamp(_cell(ws, end_row_idx,   4))

    wb.close()

    pvt    = PVTProperties(oil_shrinkage=oil_shrinkage,
                           flash_factor=flash_factor,
                           bsw=bsw)
    window = TestWindow(start=start_time, end=end_time)
    return pvt, window


# ---------------------------------------------------------------------------
# Time-series reader  (knows rows 28+, columns D–S)
# ---------------------------------------------------------------------------

def read_timeseries(filepath: str,
                    sheet_name: str = "MPFM VALIDATION DATA",
                    header_row: int = 28) -> pd.DataFrame:
    """
    Read the minute-by-minute time-series data starting at *header_row*
    (1-indexed Excel row — row 28 in the canonical layout).

    Hard-coded column mapping:
      D(4)  → timestamp
      E(5)  → sep_total_liquid   (bbl/day at test P,T)
      F(6)  → sep_gas            (mscf/day at test P,T)
      G(7)  → sep_temperature    (°F)
      H(8)  → sep_pressure       (psig)
      I(9)  → sep_gas_dp         (orifice ΔP)
      J(10) → mpfm1_oil          (stb/day)
      K(11) → mpfm1_gas          (mmscf/day)
      L(12) → mpfm1_water        (stb/day)
      M(13) → mpfm2_oil
      N(14) → mpfm2_gas
      O(15) → mpfm2_water
      P(16) → mpfm3_oil
      Q(17) → mpfm3_gas
      R(18) → mpfm3_water
      S(19) → spot_wlr           (%)

    Returns a DataFrame with a DatetimeIndex.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]

    col_map = {
        4:  "timestamp",
        5:  "sep_total_liquid",
        6:  "sep_gas",
        7:  "sep_temperature",
        8:  "sep_pressure",
        9:  "sep_gas_dp",
        10: "mpfm1_oil",
        11: "mpfm1_gas",
        12: "mpfm1_water",
        13: "mpfm2_oil",
        14: "mpfm2_gas",
        15: "mpfm2_water",
        16: "mpfm3_oil",
        17: "mpfm3_gas",
        18: "mpfm3_water",
        19: "spot_wlr",
    }

    rows = []
    for row in ws.iter_rows(min_row=header_row, max_col=19, values_only=True):
        if row[3] is None:  # col D — timestamp
            continue
        record = {name: row[col_idx - 1] for col_idx, name in col_map.items()}
        rows.append(record)

    wb.close()

    df = pd.DataFrame(rows)
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp"])
    df = df.set_index("timestamp").sort_index()

    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    return df


# ---------------------------------------------------------------------------
# CLI convenience orchestration
# ---------------------------------------------------------------------------

def analyse(filepath: str,
            sheet_name: str = "MPFM VALIDATION DATA") -> MPFMAnalysisResult:
    """Run the full MPFM validation analysis on the given spreadsheet."""
    pvt, window = read_metadata(filepath, sheet_name)
    raw  = read_timeseries(filepath, sheet_name)
    ts   = filter_test_window(raw, window)
    ts   = compute_derived_columns(ts, pvt)
    devs = compute_deviations(ts)
    comparison = build_comparison_table(ts, devs)

    return MPFMAnalysisResult(ts=ts, comparison=comparison,
                              deviations=devs, pvt=pvt, window=window)


def print_summary(result: MPFMAnalysisResult):
    """Pretty-print the analysis summary to stdout."""
    print("=" * 72)
    print("MPFM VALIDATION ANALYSIS SUMMARY")
    print("=" * 72)
    print(f"Test window : {result.window.start} → {result.window.end}")
    print(f"Data points : {len(result.ts)}")
    print(f"PVT: shrinkage={result.pvt.oil_shrinkage:.4f}, "
          f"flash={result.pvt.flash_factor:.2f} scf/stb, "
          f"BS&W={result.pvt.bsw:.4f}")
    print()

    comp = result.comparison
    fmt_pct = lambda x: f"{x*100:+.2f}%" if pd.notna(x) else "N/A"
    fmt_val = lambda x: f"{x:,.2f}"       if pd.notna(x) else "N/A"

    header = (f"{'Phase':<8} {'Unit':<10} {'MPFM Mean':>12} {'Sep Ref':>12} "
              f"{'Rel Dev':>10} {'SE(rel)':>10} {'95% CI':>22} {'Accept':>8}")
    print(header)
    print("-" * len(header))

    for _, r in comp.iterrows():
        ci     = f"[{fmt_pct(r.ci95_rel_lower)}, {fmt_pct(r.ci95_rel_upper)}]"
        accept = ("PASS" if r.within_acceptance else "FAIL") if r.acceptance_limit is not None else ""
        print(f"{r.phase:<8} {r.unit:<10} {fmt_val(r.mpfm_mean):>12} "
              f"{fmt_val(r.sep_ref_mean):>12} {fmt_pct(r.mean_rel_deviation):>10} "
              f"{fmt_pct(r.se_rel_deviation):>10} {ci:>22} {accept:>8}")
    print()


def generate_plots(result: MPFMAnalysisResult, output_dir: str):
    """Generate diagnostic plots and save to output_dir."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    out   = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    devs  = result.deviations
    phases = ["oil", "gas", "water", "liquid"]

    # Time-series overlay
    fig, axes = plt.subplots(len(phases), 1, figsize=(14, 3.5 * len(phases)), sharex=True)
    for ax, phase in zip(axes, phases):
        ax.plot(devs.index, devs[f"{phase}_mpfm"], label="MPFM",    alpha=0.8)
        ax.plot(devs.index, devs[f"{phase}_sep"],  label="Sep Ref", alpha=0.8)
        ax.set_ylabel(phase.capitalize())
        ax.legend(loc="upper right", fontsize=8)
        ax.grid(True, alpha=0.3)
    axes[0].set_title("MPFM vs Separator Reference – Time Series")
    axes[-1].set_xlabel("Time")
    plt.tight_layout()
    fig.savefig(out / "timeseries_overlay.png", dpi=150)
    plt.close(fig)

    # Deviation histograms
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    for ax, phase in zip(axes.flat, phases):
        data = devs[f"{phase}_rel_dev"].dropna() * 100
        ax.hist(data, bins=50, edgecolor="black", alpha=0.7)
        ax.axvline(data.mean(), color="red", linestyle="--",
                   label=f"Mean={data.mean():.2f}%")
        ax.set_title(f"{phase.capitalize()} Relative Deviation")
        ax.set_xlabel("Relative Deviation (%)")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)
    plt.suptitle("Distribution of Per-Timestep Relative Deviations", y=1.01)
    plt.tight_layout()
    fig.savefig(out / "deviation_histograms.png", dpi=150)
    plt.close(fig)

    print(f"Plots saved to {out}/")


def main():
    parser = argparse.ArgumentParser(
        description="MPFM Validation Analysis – compare MPFM vs separator reference"
    )
    parser.add_argument("spreadsheet", help="Path to the MPFM validation .xlsx file")
    parser.add_argument("--sheet",  default="MPFM VALIDATION DATA")
    parser.add_argument("--output", default=None)
    parser.add_argument("--plots",  action="store_true")
    parser.add_argument("--csv",    action="store_true")
    args = parser.parse_args()

    result  = analyse(args.spreadsheet, args.sheet)
    print_summary(result)

    out_dir = args.output or str(Path(args.spreadsheet).parent)

    if args.plots:
        generate_plots(result, out_dir)

    if args.csv:
        out = Path(out_dir)
        out.mkdir(parents=True, exist_ok=True)
        result.comparison.to_csv(out / "comparison_summary.csv",    index=False)
        result.deviations.to_csv(out / "deviations_timeseries.csv")
        result.ts.to_csv(        out / "filtered_timeseries.csv")
        print(f"CSV files saved to {out}/")


if __name__ == "__main__":
    main()
