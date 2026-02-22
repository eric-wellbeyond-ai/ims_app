"""
MPFM Validation Analysis Script
================================
Compares Multiphase Flow Meter (MPFM) readings against reference separator
measurements, applying PVT corrections (shrinkage factor, flash/gas evolution)
to put both on the same standard-condition basis.

Usage:
    python mpfm_analysis.py <spreadsheet_path> [--sheet SHEET_NAME]
                            [--output OUTPUT_DIR] [--plots]

The spreadsheet is expected to follow the MPFM validation layout:
  - Header metadata in rows 1-27 (PVT properties, test start/end times, etc.)
  - Time-series data from row 28 onward with columns:
      D: Timestamp
      E: Separator Total Liquid (bbl/day)
      F: Separator Gas (mscf/day)
      G: Separator Temperature (deg F)
      H: Separator Pressure (psig)
      I: Separator Gas Orifice Diff Pressure
      J-L: MPFM #1 Oil/Gas/Water (stb/day, mmscf/day, stb/day)
      M-O: MPFM #2 Oil/Gas/Water
      P-R: MPFM #3 Oil/Gas/Water
      S: Spot sample water-liquid ratio (%)
"""

import argparse
import sys
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import numpy as np
import pandas as pd
from scipy import stats


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class PVTProperties:
    """PVT conversion properties read from the spreadsheet header."""
    oil_shrinkage: float        # Bo^-1 style shrinkage (multiply stock-tank vol)
    flash_factor: float         # scf/stb – gas evolved when oil flashes to STD
    bsw: float                  # basic sediment & water fraction (0-1)


@dataclass
class MeasurementUncertainties:
    """
    Relative (fractional) uncertainties for each input channel and PVT property.
    All values are fractions, e.g. 0.05 = 5%.  Zero means no uncertainty assumed.
    """
    r_sep_liquid:    float = 0.0
    r_sep_gas:       float = 0.0
    r_mpfm_oil:      float = 0.0
    r_mpfm_gas:      float = 0.0
    r_mpfm_water:    float = 0.0
    r_bsw:           float = 0.0
    r_oil_shrinkage: float = 0.0
    r_flash_factor:  float = 0.0


@dataclass
class TestWindow:
    """Defines the valid test period."""
    start: pd.Timestamp
    end: pd.Timestamp


@dataclass
class MPFMAnalysisResult:
    """Container for all analysis outputs."""
    # Filtered time-series (only the test window)
    ts: pd.DataFrame
    # Summary comparison table (one row per phase)
    comparison: pd.DataFrame
    # Per-timestep deviation series
    deviations: pd.DataFrame
    # PVT properties used
    pvt: PVTProperties
    # Test window used
    window: TestWindow


# ---------------------------------------------------------------------------
# Reading helpers
# ---------------------------------------------------------------------------
def _cell(ws, row, col):
    """Read a single cell value from an openpyxl worksheet (1-indexed)."""
    return ws.cell(row=row, column=col).value


def read_metadata(filepath: str, sheet_name: str = "MPFM VALIDATION DATA"):
    """
    Extract PVT properties and test window from the spreadsheet header.

    Returns (PVTProperties, TestWindow).
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    # PVT properties  (row 9)
    oil_shrinkage = float(_cell(ws, 9, 14))   # N9
    flash_factor  = float(_cell(ws, 9, 15))   # O9
    bsw           = float(_cell(ws, 9, 12))   # L9

    # Test window from row indices stored in A12, B12
    start_row_idx = int(_cell(ws, 12, 1))     # A12
    end_row_idx   = int(_cell(ws, 12, 2))     # B12
    start_time = pd.Timestamp(_cell(ws, start_row_idx, 4))   # col D
    end_time   = pd.Timestamp(_cell(ws, end_row_idx, 4))

    wb.close()

    pvt = PVTProperties(oil_shrinkage=oil_shrinkage,
                        flash_factor=flash_factor,
                        bsw=bsw)
    window = TestWindow(start=start_time, end=end_time)
    return pvt, window


def read_timeseries(filepath: str,
                    sheet_name: str = "MPFM VALIDATION DATA",
                    header_row: int = 28) -> pd.DataFrame:
    """
    Read the minute-by-minute time-series data starting at *header_row*
    (1-indexed Excel row where data begins – row 28 in the sample).

    Returns a DataFrame with a DatetimeIndex and friendly column names.
    """
    # pandas header param is 0-indexed; data starts at row 28 → skiprows 0..26
    # We'll just read from the sheet directly using openpyxl for reliability.
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]

    col_map = {
        4:  "timestamp",
        5:  "sep_total_liquid",    # bbl/day at test P,T
        6:  "sep_gas",             # mscf/day at test P,T
        7:  "sep_temperature",     # deg F
        8:  "sep_pressure",        # psig
        9:  "sep_gas_dp",          # orifice diff pressure
        10: "mpfm1_oil",           # stb/day
        11: "mpfm1_gas",           # mmscf/day
        12: "mpfm1_water",         # stb/day
        13: "mpfm2_oil",
        14: "mpfm2_gas",
        15: "mpfm2_water",
        16: "mpfm3_oil",
        17: "mpfm3_gas",
        18: "mpfm3_water",
        19: "spot_wlr",            # %
    }

    rows = []
    for row in ws.iter_rows(min_row=header_row, max_col=19, values_only=True):
        ts_val = row[3]  # column D (0-indexed = 3)
        if ts_val is None:
            continue
        record = {}
        for col_idx, name in col_map.items():
            val = row[col_idx - 1]  # 0-indexed
            record[name] = val
        rows.append(record)

    wb.close()

    df = pd.DataFrame(rows)
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp"])
    df = df.set_index("timestamp").sort_index()

    # Coerce numerics
    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    return df


# ---------------------------------------------------------------------------
# Core analysis
# ---------------------------------------------------------------------------
def filter_test_window(df: pd.DataFrame, window: TestWindow) -> pd.DataFrame:
    """Return only rows within the test period (inclusive)."""
    mask = (df.index >= window.start) & (df.index <= window.end)
    filtered = df.loc[mask].copy()
    if filtered.empty:
        raise ValueError(
            f"No data found in test window {window.start} – {window.end}. "
            f"Data spans {df.index.min()} – {df.index.max()}."
        )
    return filtered


def compute_derived_columns(df: pd.DataFrame, pvt: PVTProperties) -> pd.DataFrame:
    """
    Add derived columns for both MPFM totals and separator reference
    volumes converted to standard conditions.
    """
    out = df.copy()

    # --- MPFM totals (already at standard conditions) ---
    out["mpfm_oil"]    = out["mpfm1_oil"]   + out["mpfm2_oil"]   + out["mpfm3_oil"]
    out["mpfm_water"]  = out["mpfm1_water"] + out["mpfm2_water"] + out["mpfm3_water"]
    out["mpfm_liquid"] = out["mpfm_oil"] + out["mpfm_water"]
    # Gas: MPFM reports in mmscf/day → convert to mscf/day (* 1000)
    out["mpfm_gas"]    = (out["mpfm1_gas"] + out["mpfm2_gas"] + out["mpfm3_gas"]) * 1000.0

    # --- Separator reference → standard conditions ---
    # Free water at test conditions
    out["sep_free_water"]       = out["sep_total_liquid"] * pvt.bsw
    # Oil before shrinkage (at test P,T)
    out["sep_oil_before_shrink"] = out["sep_total_liquid"] - out["sep_free_water"]
    # Oil after shrinkage → stock-tank barrels
    out["sep_oil_std"]          = out["sep_oil_before_shrink"] * pvt.oil_shrinkage
    # Total liquid at standard conditions
    out["sep_liquid_std"]       = out["sep_free_water"] + out["sep_oil_std"]
    # Flash gas (gas evolved from oil)
    out["sep_flash_gas"]        = out["sep_oil_std"] * pvt.flash_factor / 1000.0  # mscf/day
    # Total gas at standard conditions
    out["sep_gas_std"]          = out["sep_gas"] + out["sep_flash_gas"]

    # --- Water cut ---
    out["mpfm_wc"] = out["mpfm_water"] / out["mpfm_liquid"]
    out["sep_wc"]  = out["sep_free_water"] / out["sep_liquid_std"]

    # --- GOR (scf/stb) ---
    out["mpfm_gor"] = out["mpfm_gas"] * 1000.0 / out["mpfm_oil"]   # mscf→scf
    out["sep_gor"]  = out["sep_gas_std"] * 1000.0 / out["sep_oil_std"]

    return out


def compute_uncertainty_columns(
    df: pd.DataFrame,
    pvt: PVTProperties,
    unc: MeasurementUncertainties,
) -> pd.DataFrame:
    """
    Compute per-timestep 1-sigma absolute uncertainties for all derived
    quantities using first-order Gaussian error propagation.

    Returns a DataFrame with the same index as *df* containing sigma_* columns.
    All sigmas are 0.0 when the corresponding unc field is 0.0.
    """
    out = pd.DataFrame(index=df.index)

    # ------------------------------------------------------------------
    # 1. MPFM totals – sum of three independent meters, same relative unc
    # ------------------------------------------------------------------
    out["sigma_mpfm_oil"] = unc.r_mpfm_oil * np.sqrt(
        df["mpfm1_oil"] ** 2 + df["mpfm2_oil"] ** 2 + df["mpfm3_oil"] ** 2
    )
    out["sigma_mpfm_water"] = unc.r_mpfm_water * np.sqrt(
        df["mpfm1_water"] ** 2 + df["mpfm2_water"] ** 2 + df["mpfm3_water"] ** 2
    )
    out["sigma_mpfm_gas"] = unc.r_mpfm_gas * np.sqrt(
        df["mpfm1_gas"] ** 2 + df["mpfm2_gas"] ** 2 + df["mpfm3_gas"] ** 2
    ) * 1000.0  # match the *1000 unit conversion in compute_derived_columns
    out["sigma_mpfm_liquid"] = np.sqrt(
        out["sigma_mpfm_oil"] ** 2 + out["sigma_mpfm_water"] ** 2
    )

    # ------------------------------------------------------------------
    # 2. Separator free water:  sep_free_water = sep_total_liquid * bsw
    #    ∂/∂sep_liq = bsw,  ∂/∂bsw = sep_liq
    # ------------------------------------------------------------------
    sigma_sep_liq = unc.r_sep_liquid * df["sep_total_liquid"]
    sigma_bsw     = unc.r_bsw * pvt.bsw

    sigma_sep_free_water = np.sqrt(
        (pvt.bsw * sigma_sep_liq) ** 2
        + (df["sep_total_liquid"] * sigma_bsw) ** 2
    )
    out["sigma_sep_free_water"] = sigma_sep_free_water

    # ------------------------------------------------------------------
    # 3. Sep oil at standard conditions:
    #    sep_oil_std = sep_total_liquid * (1-bsw) * oil_shrinkage
    #    ∂/∂sep_liq   = (1-bsw)*shrink
    #    ∂/∂bsw       = -sep_liq*shrink   (sign irrelevant in σ²)
    #    ∂/∂shrink    = sep_liq*(1-bsw)
    # ------------------------------------------------------------------
    one_minus_bsw   = 1.0 - pvt.bsw
    sigma_shrinkage = unc.r_oil_shrinkage * pvt.oil_shrinkage

    out["sigma_sep_oil_std"] = np.sqrt(
        (pvt.oil_shrinkage * one_minus_bsw * sigma_sep_liq) ** 2
        + (pvt.oil_shrinkage * df["sep_total_liquid"] * sigma_bsw) ** 2
        + (one_minus_bsw * df["sep_total_liquid"] * sigma_shrinkage) ** 2
    )

    # ------------------------------------------------------------------
    # 4. Sep liquid at standard conditions:
    #    sep_liquid_std = sep_free_water + sep_oil_std
    # ------------------------------------------------------------------
    out["sigma_sep_liquid_std"] = np.sqrt(
        sigma_sep_free_water ** 2 + out["sigma_sep_oil_std"] ** 2
    )

    # ------------------------------------------------------------------
    # 5. Sep gas at standard conditions:
    #    sep_gas_std = sep_gas + sep_oil_std * flash_factor / 1000
    #    ∂/∂sep_gas      = 1
    #    ∂/∂sep_oil_std  = flash/1000
    #    ∂/∂flash        = sep_oil_std/1000
    # ------------------------------------------------------------------
    sigma_sep_gas = unc.r_sep_gas * df["sep_gas"]
    sigma_flash   = unc.r_flash_factor * pvt.flash_factor

    out["sigma_sep_gas_std"] = np.sqrt(
        sigma_sep_gas ** 2
        + (pvt.flash_factor / 1000.0 * out["sigma_sep_oil_std"]) ** 2
        + (df["sep_oil_std"] / 1000.0 * sigma_flash) ** 2
    )

    # ------------------------------------------------------------------
    # 6. Water cut:  wc = numerator / denominator
    #    σ_wc² = (σ_num/denom)² + (num·σ_denom/denom²)²
    # ------------------------------------------------------------------
    out["sigma_mpfm_wc"] = np.sqrt(
        (out["sigma_mpfm_water"] / df["mpfm_liquid"]) ** 2
        + (df["mpfm_water"] * out["sigma_mpfm_liquid"] / df["mpfm_liquid"] ** 2) ** 2
    )
    out["sigma_sep_wc"] = np.sqrt(
        (sigma_sep_free_water / df["sep_liquid_std"]) ** 2
        + (df["sep_free_water"] * out["sigma_sep_liquid_std"] / df["sep_liquid_std"] ** 2) ** 2
    )

    # ------------------------------------------------------------------
    # 7. GOR:  gor = gas * 1000 / oil
    # ------------------------------------------------------------------
    out["sigma_mpfm_gor"] = np.sqrt(
        (1000.0 * out["sigma_mpfm_gas"] / df["mpfm_oil"]) ** 2
        + (1000.0 * df["mpfm_gas"] * out["sigma_mpfm_oil"] / df["mpfm_oil"] ** 2) ** 2
    )
    out["sigma_sep_gor"] = np.sqrt(
        (1000.0 * out["sigma_sep_gas_std"] / df["sep_oil_std"]) ** 2
        + (1000.0 * df["sep_gas_std"] * out["sigma_sep_oil_std"] / df["sep_oil_std"] ** 2) ** 2
    )

    return out


def compute_deviations(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute per-timestep relative and absolute deviations between
    MPFM and separator reference for each phase.
    """
    phases = {
        "oil":    ("mpfm_oil",    "sep_oil_std"),
        "gas":    ("mpfm_gas",    "sep_gas_std"),
        "water":  ("mpfm_water",  "sep_free_water"),
        "liquid": ("mpfm_liquid", "sep_liquid_std"),
        "wc":     ("mpfm_wc",    "sep_wc"),
        "gor":    ("mpfm_gor",   "sep_gor"),
    }

    devs = pd.DataFrame(index=df.index)
    for phase, (mpfm_col, sep_col) in phases.items():
        devs[f"{phase}_abs_dev"]  = df[mpfm_col] - df[sep_col]
        devs[f"{phase}_rel_dev"]  = df[mpfm_col] / df[sep_col] - 1.0
        devs[f"{phase}_mpfm"]    = df[mpfm_col]
        devs[f"{phase}_sep"]     = df[sep_col]

    return devs


def build_comparison_table(
    df: pd.DataFrame,
    devs: pd.DataFrame,
    sigma_df: "pd.DataFrame | None" = None,
) -> pd.DataFrame:
    """
    Build a summary comparison table with mean values, deviations,
    standard errors, and confidence intervals.
    """
    phases = ["oil", "gas", "water", "liquid", "wc", "gor"]
    units  = {
        "oil": "STB/day", "gas": "MSCF/day", "water": "STB/day",
        "liquid": "STB/day", "wc": "fraction", "gor": "SCF/STB",
    }
    acceptance = {
        "oil": 0.05, "gas": 0.05, "water": 0.05,
        "liquid": None, "wc": None, "gor": None,
    }

    n = len(df)
    records = []
    for phase in phases:
        mpfm_mean = devs[f"{phase}_mpfm"].mean()
        sep_mean  = devs[f"{phase}_sep"].mean()
        abs_dev   = devs[f"{phase}_abs_dev"]
        rel_dev   = devs[f"{phase}_rel_dev"]

        mean_rel  = rel_dev.mean()
        std_rel   = rel_dev.std(ddof=1)
        se_rel    = std_rel / np.sqrt(n)

        mean_abs  = abs_dev.mean()
        std_abs   = abs_dev.std(ddof=1)
        se_abs    = std_abs / np.sqrt(n)

        # 95% CI on mean relative deviation
        ci95_rel  = 1.96 * se_rel

        # Propagated measurement uncertainty on means
        _sigma_mpfm_col = {
            "oil": "sigma_mpfm_oil", "gas": "sigma_mpfm_gas",
            "water": "sigma_mpfm_water", "liquid": "sigma_mpfm_liquid",
            "wc": "sigma_mpfm_wc", "gor": "sigma_mpfm_gor",
        }
        _sigma_sep_col = {
            "oil": "sigma_sep_oil_std", "gas": "sigma_sep_gas_std",
            "water": "sigma_sep_free_water", "liquid": "sigma_sep_liquid_std",
            "wc": "sigma_sep_wc", "gor": "sigma_sep_gor",
        }
        sigma_mpfm_mean = None
        sigma_sep_mean  = None
        sigma_rel_dev   = None
        if sigma_df is not None and not sigma_df.empty:
            mc = _sigma_mpfm_col.get(phase)
            sc = _sigma_sep_col.get(phase)
            if mc and mc in sigma_df.columns:
                sigma_mpfm_mean = float(sigma_df[mc].mean())
            if sc and sc in sigma_df.columns:
                sigma_sep_mean = float(sigma_df[sc].mean())
            if sigma_mpfm_mean is not None and sigma_sep_mean is not None and sep_mean != 0:
                sigma_rel_dev = float(np.sqrt(
                    (sigma_mpfm_mean / sep_mean) ** 2
                    + (mpfm_mean * sigma_sep_mean / sep_mean ** 2) ** 2
                ))

        # Paired z-test: are the datasets statistically different?
        # H0: mean(MPFM - Sep) = 0
        # With large n, use z = mean_abs_dev / SE_abs_dev
        if se_abs > 0:
            z_stat = mean_abs / se_abs
            p_value = 2.0 * (1.0 - stats.norm.cdf(abs(z_stat)))
        else:
            z_stat = 0.0
            p_value = 1.0

        rec = {
            "phase": phase,
            "unit": units[phase],
            "mpfm_mean": mpfm_mean,
            "sep_ref_mean": sep_mean,
            "mean_abs_deviation": mean_abs,
            "mean_rel_deviation": mean_rel,
            "std_rel_deviation": std_rel,
            "se_rel_deviation": se_rel,
            "ci95_rel_lower": mean_rel - ci95_rel,
            "ci95_rel_upper": mean_rel + ci95_rel,
            "std_abs_deviation": std_abs,
            "se_abs_deviation": se_abs,
            "z_statistic": z_stat,
            "p_value": p_value,
            "sigma_mpfm_mean": sigma_mpfm_mean,
            "sigma_sep_mean":  sigma_sep_mean,
            "sigma_rel_dev":   sigma_rel_dev,
            "n_samples": n,
            "acceptance_limit": acceptance.get(phase),
            "within_acceptance": (
                abs(mean_rel) <= acceptance[phase]
                if acceptance.get(phase) is not None
                else None
            ),
        }
        records.append(rec)

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------
def analyse(filepath: str,
            sheet_name: str = "MPFM VALIDATION DATA") -> MPFMAnalysisResult:
    """
    Run the full MPFM validation analysis on the given spreadsheet.

    Returns an MPFMAnalysisResult with all data and statistics.
    """
    pvt, window = read_metadata(filepath, sheet_name)
    raw = read_timeseries(filepath, sheet_name)
    ts = filter_test_window(raw, window)
    ts = compute_derived_columns(ts, pvt)
    devs = compute_deviations(ts)
    comparison = build_comparison_table(ts, devs)

    return MPFMAnalysisResult(
        ts=ts,
        comparison=comparison,
        deviations=devs,
        pvt=pvt,
        window=window,
    )


def print_summary(result: MPFMAnalysisResult):
    """Pretty-print the analysis summary to stdout."""
    print("=" * 72)
    print("MPFM VALIDATION ANALYSIS SUMMARY")
    print("=" * 72)
    print(f"Test window : {result.window.start} → {result.window.end}")
    print(f"Data points : {len(result.ts)}")
    print(f"PVT: shrinkage={result.pvt.oil_shrinkage:.4f}, "
          f"flash={result.pvt.flash_factor:.2f} scf/stb, "
          f"BS&W={result.pvt.bsw:.4f}")
    print()

    comp = result.comparison
    fmt_pct = lambda x: f"{x*100:+.2f}%" if pd.notna(x) else "N/A"
    fmt_val = lambda x: f"{x:,.2f}" if pd.notna(x) else "N/A"

    header = (f"{'Phase':<8} {'Unit':<10} {'MPFM Mean':>12} {'Sep Ref':>12} "
              f"{'Rel Dev':>10} {'SE(rel)':>10} {'95% CI':>22} "
              f"{'Accept':>8}")
    print(header)
    print("-" * len(header))

    for _, r in comp.iterrows():
        ci = f"[{fmt_pct(r.ci95_rel_lower)}, {fmt_pct(r.ci95_rel_upper)}]"
        accept = ""
        if r.acceptance_limit is not None:
            accept = "PASS" if r.within_acceptance else "FAIL"
        print(f"{r.phase:<8} {r.unit:<10} {fmt_val(r.mpfm_mean):>12} "
              f"{fmt_val(r.sep_ref_mean):>12} {fmt_pct(r.mean_rel_deviation):>10} "
              f"{fmt_pct(r.se_rel_deviation):>10} {ci:>22} {accept:>8}")

    print()


def generate_plots(result: MPFMAnalysisResult, output_dir: str):
    """Generate diagnostic plots and save to output_dir."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    devs = result.deviations

    phases = ["oil", "gas", "water", "liquid"]

    # --- Time-series overlay ---
    fig, axes = plt.subplots(len(phases), 1, figsize=(14, 3.5 * len(phases)),
                             sharex=True)
    for ax, phase in zip(axes, phases):
        ax.plot(devs.index, devs[f"{phase}_mpfm"], label="MPFM", alpha=0.8)
        ax.plot(devs.index, devs[f"{phase}_sep"],  label="Sep Ref", alpha=0.8)
        ax.set_ylabel(phase.capitalize())
        ax.legend(loc="upper right", fontsize=8)
        ax.grid(True, alpha=0.3)
    axes[0].set_title("MPFM vs Separator Reference – Time Series")
    axes[-1].set_xlabel("Time")
    plt.tight_layout()
    fig.savefig(out / "timeseries_overlay.png", dpi=150)
    plt.close(fig)

    # --- Relative deviation histograms ---
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    for ax, phase in zip(axes.flat, phases):
        data = devs[f"{phase}_rel_dev"].dropna() * 100
        ax.hist(data, bins=50, edgecolor="black", alpha=0.7)
        ax.axvline(data.mean(), color="red", linestyle="--",
                   label=f"Mean={data.mean():.2f}%")
        ax.set_title(f"{phase.capitalize()} Relative Deviation")
        ax.set_xlabel("Relative Deviation (%)")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)
    plt.suptitle("Distribution of Per-Timestep Relative Deviations", y=1.01)
    plt.tight_layout()
    fig.savefig(out / "deviation_histograms.png", dpi=150)
    plt.close(fig)

    # --- Deviation time series ---
    fig, axes = plt.subplots(len(phases), 1, figsize=(14, 3.5 * len(phases)),
                             sharex=True)
    for ax, phase in zip(axes, phases):
        pct = devs[f"{phase}_rel_dev"] * 100
        ax.plot(devs.index, pct, alpha=0.7, linewidth=0.8)
        ax.axhline(0, color="black", linewidth=0.5)
        ax.axhline(5, color="red", linestyle="--", linewidth=0.5, alpha=0.5)
        ax.axhline(-5, color="red", linestyle="--", linewidth=0.5, alpha=0.5)
        ax.set_ylabel(f"{phase.capitalize()} (%)")
        ax.grid(True, alpha=0.3)
    axes[0].set_title("Relative Deviation Over Time (±5% limits shown)")
    axes[-1].set_xlabel("Time")
    plt.tight_layout()
    fig.savefig(out / "deviation_timeseries.png", dpi=150)
    plt.close(fig)

    # --- Cross-plots ---
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    for ax, phase in zip(axes.flat, phases):
        x = devs[f"{phase}_sep"]
        y = devs[f"{phase}_mpfm"]
        ax.scatter(x, y, alpha=0.15, s=8)
        lims = [min(x.min(), y.min()), max(x.max(), y.max())]
        ax.plot(lims, lims, "r--", linewidth=1, label="1:1 line")
        ax.set_xlabel("Separator Reference")
        ax.set_ylabel("MPFM")
        ax.set_title(phase.capitalize())
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)
    plt.suptitle("Cross-Plot: MPFM vs Separator Reference", y=1.01)
    plt.tight_layout()
    fig.savefig(out / "crossplots.png", dpi=150)
    plt.close(fig)

    print(f"Plots saved to {out}/")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="MPFM Validation Analysis – compare MPFM vs separator reference"
    )
    parser.add_argument("spreadsheet", help="Path to the MPFM validation .xlsx file")
    parser.add_argument("--sheet", default="MPFM VALIDATION DATA",
                        help="Sheet name (default: 'MPFM VALIDATION DATA')")
    parser.add_argument("--output", default=None,
                        help="Output directory for plots (default: same dir as input)")
    parser.add_argument("--plots", action="store_true",
                        help="Generate diagnostic plots")
    parser.add_argument("--csv", action="store_true",
                        help="Export results to CSV files")
    args = parser.parse_args()

    result = analyse(args.spreadsheet, args.sheet)
    print_summary(result)

    out_dir = args.output or str(Path(args.spreadsheet).parent)

    if args.plots:
        generate_plots(result, out_dir)

    if args.csv:
        out = Path(out_dir)
        out.mkdir(parents=True, exist_ok=True)
        result.comparison.to_csv(out / "comparison_summary.csv", index=False)
        result.deviations.to_csv(out / "deviations_timeseries.csv")
        result.ts.to_csv(out / "filtered_timeseries.csv")
        print(f"CSV files saved to {out}/")


if __name__ == "__main__":
    main()
